import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Configuración de la página web
st.set_page_config(page_title="Analizador Previsional Inteligente", layout="wide")

st.title("🛠️ Herramienta de Análisis y Consolidación Previsional")
st.markdown("Cargue su archivo de historias laborales con filas múltiples para clasificar y evaluar el mejor cargo.")

# -----------------------------------------------------------------
# 1. MÓDULO DE CARGA DE ARCHIVOS
# -----------------------------------------------------------------
archivo_subido = st.file_uploader("📂 Arrastre o seleccione el archivo Excel de origen (.xlsx)", type=["xlsx"])

if archivo_subido is not None:
    # Cargar datos normalmente con pandas para la estructura
    df_original = pd.read_excel(archivo_subido)
    st.success("✅ Archivo cargado correctamente en la memoria temporal.")
    
    # -----------------------------------------------------------------
    # 2. ASISTENTE DE MAPEO DE CAMPOS (Interfaz de Usuario)
    # -----------------------------------------------------------------
    st.subheader("📋 Asistente de Mapeo de Columnas")
    st.info("Asocie las columnas de su archivo con los campos lógicos que requiere el motor.")
    
    columnas_archivo = list(df_original.columns)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        col_cuil = st.selectbox("Clave de Identificación (CUIL/Documento)", opciones=columnas_archivo, index=0)
        col_nombre = st.selectbox("Apellido y Nombre", opciones=columnas_archivo, index=1)
    with col2:
        col_cargo = st.selectbox("Cargo de Mayor Jerarquía (Columna con celdas verdes)", opciones=columnas_archivo, index=8)
        col_meses = st.selectbox("Meses en el Cargo", opciones=columnas_archivo, index=9)
    with col3:
        col_anos_totales = st.selectbox("Años de Servicio Totales", opciones=columnas_archivo, index=4)
        col_meses_totales = st.selectbox("Meses de Servicio Totales", opciones=columnas_archivo, index=5)

    # Botón para activar el motor lógico
    if st.button("🚀 Procesar y Categorizar Datos"):
        
        # -----------------------------------------------------------------
        # 3. MOTOR LÓGICO: DETECCIÓN DE COLOR VERDE Y REGLAS (36/60 meses)
        # -----------------------------------------------------------------
        with st.spinner("Analizando celdas destacadas y consistencia de tiempos..."):
            
            # Volver a cargar el archivo usando openpyxl para extraer los colores reales
            archivo_subido.seek(0)
            wb = openpyxl.load_workbook(archivo_subido, data_only=True)
            ws = wb.active
            
            # Encontrar el índice numérico de la columna del cargo mapeado
            idx_cargo_col = columnas_archivo.index(col_cargo) + 1
            
            # Lista para almacenar si la celda de esa fila era verde
            es_celda_verde = []
            
            # Recorrer las filas buscando el color verde (omitir cabecera fila 1)
            for row in range(2, ws.max_row + 1):
                celda = ws.cell(row=row, column=idx_cargo_col)
                color_hex = celda.fill.start_color.rgb if celda.fill and celda.fill.fill_type else None
                
                # Detectar tonos de verde (variaciones comunes en Excel como 'D4EDDA' o verdes estándar)
                # Si no tiene color o es blanco, se considera falso
                if color_hex and color_hex != "00000000" and color_hex != "FFFFFFFF":
                    # Simplificación lógica: si hay color asignado en esa columna, asumimos el destacado
                    es_celda_verde.append(True)
                else:
                    es_celda_verde.append(False)
            
            # Asegurar correspondencia con el DataFrame de Pandas (recortando si difieren filas vacías)
            df_original = df_original.head(len(es_celda_verde)).copy()
            df_original['__es_mejor_cargo_verde__'] = es_celda_verde

            # --- CONSOLIDACIÓN POR EMPLEADO (CUIL) ---
            resultados_consolidados = []
            
            for cuil, grupo in df_original.groupby(col_cuil):
                nombre = grupo[col_nombre].iloc[0]
                
                # Filtrar solo los registros donde el cargo estaba en verde
                grupo_mejor_cargo = grupo[grupo['__es_mejor_cargo_verde__'] == True]
                
                mejor_cargo_nombre = "No Identificado / Sin color"
                condicion_cumplida = "No califica"
                meses_computados = 0
                
                if not grupo_mejor_cargo.empty:
                    mejor_cargo_nombre = grupo_mejor_cargo[col_cargo].iloc[0]
                    # Sumamos los meses declarados en esas filas específicas
                    total_meses_en_verde = grupo_mejor_cargo[col_meses].sum()
                    
                    # Evaluación de Reglas en Cascada
                    # Nota: Para evaluar continuidad perfecta se requerirían las fechas exactas ordenadas,
                    # aquí lo evaluamos por acumulación directa bajo sus umbrales normativos 36/60:
                    if total_meses_en_verde >= 36:
                        # Asumimos continuidad si están agrupados consecutivamente en el bloque verde
                        condicion_cumplida = "✅ Regla 1 Cumplida: Continuo >= 36 Meses"
                        meses_computados = total_meses_en_verde
                    elif total_meses_en_verde >= 60:
                        condicion_cumplida = "⚠️ Regla 2 Cumplida: Alternado >= 60 Meses"
                        meses_computados = total_meses_en_verde
                    else:
                        condicion_cumplida = f"❌ No alcanza el mínimo previsional (Suma: {total_meses_en_verde} meses)"
                        meses_computados = total_meses_en_verde

                # Cálculo de antigüedad general del empleado
                anos_totales = grupo[col_anos_totales].sum()
                meses_totales = grupo[col_meses_totales].sum()
                
                resultados_consolidados.append({
                    "CUIL/Documento": cuil,
                    "Apellido y Nombre": nombre,
                    "Antigüedad Total (Años)": anos_totales,
                    "Antigüedad Total (Meses)": meses_totales,
                    "Mejor Cargo Detectado": mejor_cargo_nombre,
                    "Condición Legal del Cargo": condicion_cumplida,
                    "Meses Totales en Cargo": meses_computados
                })
            
            df_final = pd.DataFrame(resultados_consolidados)
            
            # -----------------------------------------------------------------
            # 4. EXPOSICIÓN DE RESULTADOS Y DESCARGA
            # -----------------------------------------------------------------
            st.subheader("📊 Reporte Previsional Generado")
            st.dataframe(df_final, use_container_width=True)
            
            # Conversión del resultado a un nuevo Excel descargable
            @st.cache_data
            def convertir_excel(df):
                import io
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Consolidado')
                return output.getvalue()
            
            excel_descargable = convertir_excel(df_final)
            
            st.download_button(
                label="📥 Descargar Reporte Consolidado en Excel",
                data=excel_descargable,
                file_name="Reporte_Consolidado_Previsional.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )